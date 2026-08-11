import openpyxl, json
from datetime import datetime, time
WB='/root/.claude/uploads/34cce73d-0c23-533e-95a7-ef3cccabda66/316667f7-B20x50__CL_Beta_plan_30d_2210_8.xlsx'
wb=openpyxl.load_workbook(WB, data_only=True)
ws=wb['Entrada de Datos']
T=[]
for r in range(4, ws.max_row+1):
    sym=ws.cell(r,4).value
    fecha=ws.cell(r,5).value
    if not sym or not isinstance(fecha,datetime): continue
    def g(c): return ws.cell(r,c).value
    T.append(dict(
        n=len(T)+1, bloque=g(2), sym=str(sym).strip(),
        fecha=fecha.strftime('%Y-%m-%d'),
        hora=g(6).strftime('%H:%M') if isinstance(g(6),time) else None,
        dir=(g(7) or '').strip(), contratos=g(8), patron=(g(9) or '').strip(),
        l1=g(10), l2=g(11), l3=g(12),
        f_sal=g(13).strftime('%Y-%m-%d') if isinstance(g(13),datetime) else None,
        h_sal=g(14).strftime('%H:%M') if isinstance(g(14),time) else None,
        mfe=g(15), mae=g(16),
        disc=(g(18) or '').strip(), motivo=(g(19) or '').strip(),
        dur=g(20), ticks=g(21), ticks_com=g(22),
        usd=g(23), acum=g(24), riesgo=g(25), evol=g(26)))
json.dump(T, open('beta.json','w'), indent=0)
print('operaciones:', len(T))
print('rango:', T[0]['fecha'], '->', T[-1]['fecha'])
print('MFE cargados:', sum(1 for t in T if t['mfe'] is not None))
print('MAE cargados:', sum(1 for t in T if t['mae'] is not None))
print('simbolos:', {t['sym'] for t in T})
print('patrones:', {t['patron'] for t in T})
print('contratos:', sorted({t['contratos'] for t in T}))
print('lotes 2/3 usados:', sum(1 for t in T if t['l2'] is not None), sum(1 for t in T if t['l3'] is not None))
print('capital final:', T[-1]['acum'])
